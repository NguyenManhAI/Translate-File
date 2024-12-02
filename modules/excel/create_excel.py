from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.drawing.image import Image as ExcelImage

def create_translated_excel(input_excel_path, translated_data, output_excel_path):
    new_wb = load_workbook(input_excel_path)

    # Xóa sheet mặc định trống mà openpyxl tạo ra
    # default_sheet = new_wb["Sheet"]
    # new_wb.remove(default_sheet)

    for sheet_name, sheet_info in translated_data.items():
        # new_sheet = new_wb.create_sheet(title=sheet_name)
        new_sheet = new_wb[sheet_name]

        # # Thêm ảnh vào file excel
        # images_sheet_data = images_data[sheet_name]
        # for img_info in images_sheet_data:
        #     # Đọc hình ảnh từ đường dẫn
        #     img_path = img_info["path"]
        #     position = img_info["position"]
            
        #     # Load hình ảnh
        #     img = ExcelImage(img_path)

        #     # Chèn hình ảnh vào vị trí tương ứng
        #     new_sheet.add_image(img, position)

        for row_idx, row in enumerate(sheet_info, 1):  # Excel index starts at 1
            for col_idx, cell_info in enumerate(row, 1):
                cell = new_sheet.cell(row=row_idx, column=col_idx, value=cell_info["value"])

                # Khôi phục định dạng từng thuộc tính cho ô

                if cell_info["font"]:
                    cell.font = Font(
                        name=cell_info["font"].name,
                        size=cell_info["font"].size,
                        bold=cell_info["font"].bold,
                        italic=cell_info["font"].italic,
                        color=cell_info["font"].color
                    )
                
                if cell_info["fill"]:
                    cell.fill = PatternFill(
                        fill_type=cell_info["fill"].fill_type,
                        fgColor=cell_info["fill"].fgColor,
                        bgColor=cell_info["fill"].bgColor
                    )

                if cell_info["alignment"]:
                    cell.alignment = Alignment(
                        horizontal=cell_info["alignment"].horizontal,
                        vertical=cell_info["alignment"].vertical,
                        wrap_text=cell_info["alignment"].wrap_text
                    )

                if cell_info["border"]:
                    cell.border = Border(
                        left=cell_info["border"].left,
                        right=cell_info["border"].right,
                        top=cell_info["border"].top,
                        bottom=cell_info["border"].bottom
                    )

    # Lưu file Excel mới
    new_wb.save(output_excel_path)