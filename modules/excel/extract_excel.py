import openpyxl
import openai
import os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from PIL import Image

def extract_images_from_sheet(sheet):
    # Tạo thư mục nếu chưa tồn tại
    output_folder = "tmp"
    Path(output_folder).mkdir(parents=True, exist_ok=True)

    images_info = []

    # Duyệt qua tất cả hình ảnh trong sheet
    for i, drawing in enumerate(sheet._images, start=1):
        img_path = os.path.join(output_folder, f"{sheet.title}_image_{i}.png")
        image_data = drawing.ref
        image_data.seek(0)  # Đặt con trỏ về đầu đối tượng BytesIO để đọc dữ liệu
        image = Image.open(image_data)
        image.save(img_path)
        position = drawing.anchor

        images_info.append({"path": img_path, "position": position})

    return images_info

def extract_text_image_with_excel(excel_file, translate, src_lang, trg_lang):
    wb = openpyxl.load_workbook(excel_file)
    sheets_data = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_info = []

        for row in sheet.iter_rows():
            row_info = []
            for cell in row:
                # Lưu trữ văn bản và định dạng của từng ô
                cell_info = {
                    "value": translate(cell.value, src_lang, trg_lang) if cell.value else None,
                    # "value": cell.value,
                    "font": cell.font,
                    "fill": cell.fill,
                    "alignment": cell.alignment,
                    "border": cell.border,
                }
                row_info.append(cell_info)
            sheet_info.append(row_info)

        sheets_data[sheet_name] = sheet_info

    # # lấy thông tin ảnh
    # sheets_image_data = {}
    # for sheet_name in wb.sheetnames:
    #     sheet = wb[sheet_name]
    #     sheet_images_info = extract_images_from_sheet(sheet)
    #     sheets_image_data[sheet_name] = sheet_images_info

    return sheets_data #, sheets_image_data
